from __future__ import annotations

from copy import copy
from dataclasses import dataclass, field
from pathlib import Path
import re
from typing import Literal

import pandas as pd
from openpyxl.utils import get_column_letter


COMPARE_REQUIRED_COLUMNS = {
    "Butikkområde",
    "Butikk A",
    "Omsetning Kr",
    "Butikk B",
    "Omsetning Kr.1",
}

BRUTTO_REQUIRED_COLUMNS = {
    "Vare",
    "Brutto Kr",
    "Brutto %",
}


PRODUCT_CODE_PATTERN = re.compile(r"-\s*(\d{6,})\s+\S+$")
STORE_NAME_PATTERN = re.compile(r"^\s*(\d+)\s*-\s*(.+?)\s*$")


@dataclass(slots=True)
class CategoryTotals:
    store_a_sales: float
    store_b_sales: float


@dataclass(slots=True)
class CategoryAnalysisResult:
    category_name: str
    source_file: Path
    store_a_name: str
    store_b_name: str
    normalization_mode: str
    totals_used: CategoryTotals
    actual_category_totals: CategoryTotals
    scale_factor_b_to_a: float
    total_rows: int
    matched_rows: int
    eligible_rows: int
    positive_potential_rows: int
    warnings: list[str] = field(default_factory=list)
    table: pd.DataFrame = field(default_factory=pd.DataFrame)


@dataclass(slots=True)
class AnalysisResult:
    store_a_name: str
    store_b_name: str
    normalization_mode: str
    minimum_gross_percent: float
    top_n: int
    summary: pd.DataFrame
    category_results: list[CategoryAnalysisResult]
    brutto_base_rows: int


class AnalysisError(Exception):
    """Raised when an input file is missing required data."""


def extract_product_code(value: object) -> str | None:
    if pd.isna(value):
        return None

    text = str(value).strip()
    match = PRODUCT_CODE_PATTERN.search(text)
    if match:
        return match.group(1)

    fallback_matches = re.findall(r"(\d{6,})", text)
    return fallback_matches[-1] if fallback_matches else None


def normalize_percent(value: object) -> float | None:
    if pd.isna(value):
        return None

    try:
        numeric_value = float(value)
    except (TypeError, ValueError):
        return None

    if numeric_value <= 1.5:
        return numeric_value * 100
    return numeric_value


def parse_money_series(series: pd.Series) -> pd.Series:
    return pd.to_numeric(series, errors="coerce").fillna(0.0)


def format_percentage(value: float) -> str:
    return f"{value:.2f} %"


def validate_columns(dataframe: pd.DataFrame, required_columns: set[str], label: str) -> None:
    missing = required_columns - set(dataframe.columns)
    if missing:
        missing_list = ", ".join(sorted(missing))
        raise AnalysisError(f"{label} mangler disse kolonnene: {missing_list}")


def infer_store_names(dataframe: pd.DataFrame) -> tuple[str, str]:
    store_a_id = infer_store_id(dataframe.get("Butikk A"))
    store_b_id = infer_store_id(dataframe.get("Butikk B"))

    store_a_name = find_store_name_from_cells(dataframe, store_a_id) or fallback_store_name("Butikk A", store_a_id)
    store_b_name = find_store_name_from_cells(dataframe, store_b_id) or fallback_store_name("Butikk B", store_b_id)
    return store_a_name, store_b_name


def infer_store_id(series: pd.Series | None) -> str | None:
    if series is None:
        return None

    numeric_values = pd.to_numeric(series, errors="coerce").dropna()
    if numeric_values.empty:
        return None
    first_value = int(numeric_values.iloc[0])
    return str(first_value)


def find_store_name_from_cells(dataframe: pd.DataFrame, store_id: str | None) -> str | None:
    if not store_id:
        return None

    for value in dataframe.to_numpy().ravel():
        if pd.isna(value):
            continue
        if not isinstance(value, str):
            continue

        match = STORE_NAME_PATTERN.match(value)
        if match and match.group(1) == store_id:
            return match.group(2).strip()

    return None


def fallback_store_name(prefix: str, store_id: str | None) -> str:
    if store_id:
        return f"{prefix} ({store_id})"
    return prefix


def category_name_from_path(path: Path) -> str:
    name = path.stem
    if name.lower().startswith("sammenlign"):
        name = name[len("Sammenlign") :].strip(" _-")
    return name or path.stem


def load_brutto_file(path: Path) -> pd.DataFrame:
    dataframe = pd.read_excel(path)
    validate_columns(dataframe, BRUTTO_REQUIRED_COLUMNS, f"Bruttofila {path.name}")

    brutto = dataframe.copy()
    brutto["Tallkode"] = brutto["Vare"].map(extract_product_code)
    brutto["Brutto % norm"] = brutto["Brutto %"].map(normalize_percent)
    brutto["Brutto Kr"] = parse_money_series(brutto["Brutto Kr"])
    brutto["Omsetning Kr"] = parse_money_series(brutto.get("Omsetning Kr", pd.Series(dtype=float)))

    brutto = brutto.dropna(subset=["Tallkode", "Brutto % norm"]).copy()
    brutto = brutto.sort_values(by="Omsetning Kr", ascending=False)
    brutto = brutto.drop_duplicates(subset=["Tallkode"], keep="first")

    return brutto[["Tallkode", "Vare", "Brutto % norm", "Brutto Kr"]].rename(
        columns={
            "Vare": "Vare i bruttofil",
            "Brutto % norm": "Brutto %",
        }
    )


def load_compare_file(path: Path) -> tuple[pd.DataFrame, str, str]:
    dataframe = pd.read_excel(path)
    validate_columns(dataframe, COMPARE_REQUIRED_COLUMNS, f"Sammenligningsfila {path.name}")

    store_a_name, store_b_name = infer_store_names(dataframe)

    compare = dataframe.copy()
    compare["Vare"] = compare["Butikkområde"].astype(str).str.strip()
    compare["Tallkode"] = compare["Vare"].map(extract_product_code)
    compare["Omsetning butikk A"] = parse_money_series(compare["Omsetning Kr"])
    compare["Omsetning butikk B"] = parse_money_series(compare["Omsetning Kr.1"])
    compare["Differanse rå"] = compare["Omsetning butikk B"] - compare["Omsetning butikk A"]

    compare = compare.dropna(subset=["Tallkode"]).copy()
    return compare, store_a_name, store_b_name


def load_total_file(path: Path) -> tuple[CategoryTotals, str, str]:
    dataframe = pd.read_excel(path)
    validate_columns(dataframe, COMPARE_REQUIRED_COLUMNS, f"Totalfila {path.name}")

    store_a_name, store_b_name = infer_store_names(dataframe)
    total_a = parse_money_series(dataframe["Omsetning Kr"]).sum()
    total_b = parse_money_series(dataframe["Omsetning Kr.1"]).sum()
    return CategoryTotals(total_a, total_b), store_a_name, store_b_name


def analyze_files(
    compare_files: list[str | Path],
    brutto_file: str | Path,
    *,
    total_file: str | Path | None = None,
    minimum_gross_percent: float = 18.0,
    top_n: int = 10,
    normalization_mode: Literal["category", "total-file"] = "category",
) -> AnalysisResult:
    if not compare_files:
        raise AnalysisError("Velg minst én sammenligningsfil.")

    brutto_path = Path(brutto_file)
    brutto_lookup = load_brutto_file(brutto_path)
    brutto_base_rows = len(brutto_lookup)

    total_basis: CategoryTotals | None = None
    total_store_names: tuple[str, str] | None = None
    if normalization_mode == "total-file":
        if not total_file:
            raise AnalysisError("Normalisering mot totalfil er valgt, men totalfil mangler.")
        totals, store_a_name, store_b_name = load_total_file(Path(total_file))
        total_basis = totals
        total_store_names = (store_a_name, store_b_name)

    category_results: list[CategoryAnalysisResult] = []
    summary_rows: list[dict[str, object]] = []

    preferred_store_names: tuple[str, str] | None = total_store_names

    for compare_file in compare_files:
        compare_path = Path(compare_file)
        compare_data, store_a_name, store_b_name = load_compare_file(compare_path)

        if preferred_store_names is None:
            preferred_store_names = (store_a_name, store_b_name)

        actual_totals = CategoryTotals(
            store_a_sales=float(compare_data["Omsetning butikk A"].sum()),
            store_b_sales=float(compare_data["Omsetning butikk B"].sum()),
        )

        basis_totals = total_basis or actual_totals
        if basis_totals.store_b_sales <= 0:
            raise AnalysisError(f"Butikk B har 0 i total omsetning for {compare_path.name}.")

        scale_factor_b_to_a = basis_totals.store_a_sales / basis_totals.store_b_sales

        merged = compare_data.merge(brutto_lookup, on="Tallkode", how="left")
        total_rows = len(merged)
        matched_rows = int(merged["Brutto %"].notna().sum())

        eligible = merged.loc[merged["Brutto %"].ge(minimum_gross_percent)].copy()
        eligible_rows = len(eligible)

        eligible["Omsetning butikk B justert"] = eligible["Omsetning butikk B"] * scale_factor_b_to_a
        eligible["Potensial Bakklandet"] = eligible["Omsetning butikk B justert"] - eligible["Omsetning butikk A"]
        eligible["Potensiell brutto kr"] = eligible["Potensial Bakklandet"] * eligible["Brutto %"].fillna(0) / 100
        eligible["Potensial andel av A"] = eligible["Potensial Bakklandet"] / basis_totals.store_a_sales * 100
        eligible["Butikk B andel av grunnlag"] = eligible["Omsetning butikk B"] / basis_totals.store_b_sales * 100

        positive = eligible.loc[eligible["Potensial Bakklandet"] > 0].copy()
        positive = positive.sort_values(
            by=["Potensial Bakklandet", "Omsetning butikk B"],
            ascending=[False, False],
        )
        positive_potential_rows = len(positive)

        result_table = positive.head(top_n).copy()
        result_table["Kategori"] = category_name_from_path(compare_path)
        result_table["Filnavn"] = compare_path.name
        result_table = result_table[
            [
                "Kategori",
                "Vare",
                "Tallkode",
                "Omsetning butikk A",
                "Omsetning butikk B",
                "Omsetning butikk B justert",
                "Potensial Bakklandet",
                "Differanse rå",
                "Brutto %",
                "Potensiell brutto kr",
                "Brutto Kr",
                "Butikk B andel av grunnlag",
                "Potensial andel av A",
            ]
        ].rename(
            columns={
                "Omsetning butikk A": store_a_name,
                "Omsetning butikk B": store_b_name,
                "Omsetning butikk B justert": f"{store_b_name} justert",
                "Potensial Bakklandet": f"Potensial i {store_a_name}",
                "Differanse rå": "Rå differanse",
            }
        )

        warnings: list[str] = []
        if matched_rows == 0:
            warnings.append("Ingen varer i denne fila fant match i bruttofila.")
        else:
            match_rate = matched_rows / max(total_rows, 1)
            if match_rate < 0.25:
                warnings.append(
                    f"Bruttofila matchet bare {matched_rows} av {total_rows} varer i denne varegruppa."
                )
        if eligible_rows == 0:
            warnings.append(
                f"Ingen matchede varer i denne varegruppa hadde brutto over {format_percentage(minimum_gross_percent)}."
            )
        elif positive_potential_rows < top_n:
            warnings.append(
                f"Fant bare {positive_potential_rows} varer med positivt potensial over bruttofilteret."
            )

        summary_rows.append(
            {
                "Varegruppe": category_name_from_path(compare_path),
                "Filnavn": compare_path.name,
                "Normalisering": "Totalfil" if total_basis else "Varegruppe",
                "Total omsetning A brukt": basis_totals.store_a_sales,
                "Total omsetning B brukt": basis_totals.store_b_sales,
                "Skaleringsfaktor B->A": scale_factor_b_to_a,
                "Varer i fila": total_rows,
                "Matchet bruttofila": matched_rows,
                "Over bruttofilter": eligible_rows,
                "Positivt potensial": positive_potential_rows,
                "Advarsler": " | ".join(warnings),
            }
        )

        category_results.append(
            CategoryAnalysisResult(
                category_name=category_name_from_path(compare_path),
                source_file=compare_path,
                store_a_name=store_a_name,
                store_b_name=store_b_name,
                normalization_mode="total-file" if total_basis else "category",
                totals_used=basis_totals,
                actual_category_totals=actual_totals,
                scale_factor_b_to_a=scale_factor_b_to_a,
                total_rows=total_rows,
                matched_rows=matched_rows,
                eligible_rows=eligible_rows,
                positive_potential_rows=positive_potential_rows,
                warnings=warnings,
                table=result_table,
            )
        )

    if preferred_store_names is None:
        preferred_store_names = ("Butikk A", "Butikk B")

    summary = pd.DataFrame(summary_rows)

    return AnalysisResult(
        store_a_name=preferred_store_names[0],
        store_b_name=preferred_store_names[1],
        normalization_mode=normalization_mode,
        minimum_gross_percent=minimum_gross_percent,
        top_n=top_n,
        summary=summary,
        category_results=category_results,
        brutto_base_rows=brutto_base_rows,
    )


def export_analysis(result: AnalysisResult, output_path: str | Path) -> Path:
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        result.summary.to_excel(writer, sheet_name="Oppsummering", index=False)

        for category_result in result.category_results:
            sheet_name = safe_sheet_name(category_result.category_name)
            export_table = category_result.table.copy()
            export_table.to_excel(writer, sheet_name=sheet_name, index=False, startrow=4)

            worksheet = writer.sheets[sheet_name]
            worksheet["A1"] = f"Varegruppe: {category_result.category_name}"
            worksheet["A2"] = (
                f"Normalisering: {'Totalfil' if category_result.normalization_mode == 'total-file' else 'Per varegruppe'}"
            )
            worksheet["A3"] = (
                f"Grunnlag A/B: {category_result.totals_used.store_a_sales:.2f} / "
                f"{category_result.totals_used.store_b_sales:.2f}"
            )
            worksheet["A4"] = f"Skaleringsfaktor B->A: {category_result.scale_factor_b_to_a:.6f}"
            worksheet.freeze_panes = "A6"
            if not export_table.empty:
                for cell in worksheet[5]:
                    new_font = copy(cell.font)
                    new_font.bold = True
                    cell.font = new_font

        workbook = writer.book
        for sheet_name in workbook.sheetnames:
            autosize_worksheet(workbook[sheet_name])

    return output


def safe_sheet_name(name: str) -> str:
    cleaned = re.sub(r"[:\\\\/*?\\[\\]]", " ", name).strip()
    return cleaned[:31] or "Ark"


def autosize_worksheet(worksheet) -> None:
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            if cell.value is None:
                continue
            max_length = max(max_length, len(str(cell.value)))
        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 40)
